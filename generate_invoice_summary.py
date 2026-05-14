"""
根据对账单生成开票汇总 - 单张发票版本
对于井座、闸阀方井等产品，按高度转换为米数汇总
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict
from datetime import datetime

DEFAULT_SOURCE_FILE = "工作簿2.xlsx"


def resolve_source_files():
    """优先使用命令行传入的文件；未传入时默认只处理工作簿2.xlsx。"""
    cli_files = sys.argv[1:]
    if cli_files:
        return cli_files
    return [DEFAULT_SOURCE_FILE]


SOURCE_FILES = resolve_source_files()


def normalize_text(value):
    """标准化单元格文本，便于匹配表头。"""
    if value is None:
        return ''
    return (
        str(value)
        .replace('\n', '')
        .replace(' ', '')
        .replace('（', '(')
        .replace('）', ')')
        .strip()
    )


def find_header_row(ws):
    """在工作表前几行中定位表头行。"""
    search_end_row = min(ws.max_row, 15)
    for row_idx in range(1, search_end_row + 1):
        values = [normalize_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        if '销售日期' in values and '商品名称' in values and '数量' in values:
            return row_idx
    return None


def build_column_map(ws, header_row):
    """根据表头行构建字段到列号的映射。"""
    alias_map = {
        'date': {'销售日期'},
        'name': {'商品名称'},
        'spec': {'规格型号'},
        'unit': {'单位'},
        'qty': {'数量'},
        'price': {'单价(元)'},
        'amount': {'金额(元)'},
        'invoice_no': {'销售单据号'},
    }

    column_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_text(ws.cell(header_row, col_idx).value)
        for field_name, aliases in alias_map.items():
            if header in aliases:
                column_map[field_name] = col_idx

    required_fields = {'date', 'name', 'spec', 'unit', 'qty', 'price', 'amount'}
    if not required_fields.issubset(column_map):
        return None

    return column_map


def find_prefixed_value(ws, prefixes, search_end_row):
    """在工作表前几行查找指定前缀的信息。"""
    for row_idx in range(1, search_end_row + 1):
        value = ws.cell(row_idx, 1).value
        if value is None:
            continue
        text = str(value).strip()
        for prefix in prefixes:
            if text.startswith(prefix):
                return text
    return ''

def extract_height_in_meters(spec, name):
    """
    从规格中提取高度（最后一个数字），转换为米数
    比如 1000*150*1100 -> 1100mm -> 1.1米
    """
    if spec is None or spec == '':
        return None
    
    # 对于井座、闸阀方井等产品，规格为 宽*深*高 格式
    if '井座' in str(name) or '闸阀方井' in str(name):
        parts = str(spec).split('*')
        if len(parts) >= 3:
            try:
                height_mm = float(parts[-1])
                height_m = height_mm / 1000  # 转换为米
                return height_m
            except:
                return None
    return None

# 解析所有有效工作表中的开票数据，合并为单张发票
all_items = []
total_amount = 0
first_date = None
processed_sheet_names = []
processed_source_files = []
customer_infos = []
address_infos = []

for invoice_file in SOURCE_FILES:
    if not os.path.exists(invoice_file):
        continue

    wb = openpyxl.load_workbook(invoice_file, data_only=True)
    file_has_data = False

    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        if header_row is None:
            continue

        column_map = build_column_map(ws, header_row)
        if column_map is None:
            continue

        file_has_data = True
        processed_sheet_names.append(f"{invoice_file} / {ws.title}")

        search_end_row = min(header_row, 10)
        customer_info = find_prefixed_value(ws, ['收货单位：', '客户：'], search_end_row)
        address_info = find_prefixed_value(ws, ['送货地址：'], search_end_row)

        if customer_info and customer_info not in customer_infos:
            customer_infos.append(customer_info)
        if address_info and address_info not in address_infos:
            address_infos.append(address_info)

        for i in range(header_row + 1, ws.max_row + 1):
            date = ws.cell(i, column_map['date']).value
            name = ws.cell(i, column_map['name']).value
            spec = ws.cell(i, column_map['spec']).value
            unit = ws.cell(i, column_map['unit']).value
            qty = ws.cell(i, column_map['qty']).value
            price = ws.cell(i, column_map['price']).value
            amount = ws.cell(i, column_map['amount']).value
            invoice_no = ws.cell(i, column_map['invoice_no']).value if 'invoice_no' in column_map else None

            # 跳过空行
            if name is None and spec is None and qty is None and amount is None:
                continue

            if normalize_text(name) == '':
                continue

            # 跳过汇总行
            if name is None or '合计' in str(name):
                continue

            if amount is not None:
                amount_float = float(amount)
            else:
                amount_float = 0

            if first_date is None and date:
                first_date = date

            # 检查是否需要按高度转换
            height_m = extract_height_in_meters(spec, name)

            if height_m is not None:
                # 按高度汇总（数量 × 高度（米））
                qty_float = float(qty) if qty else 1
                total_meters = qty_float * height_m
                all_items.append({
                    'name': name,
                    'spec': spec if spec else '',
                    'unit': '米',
                    'qty': total_meters,  # 数量 × 高度 = 总米数
                    'price': price,
                    'amount': amount_float,
                    'original_qty': qty,
                    'original_unit': unit,
                    'invoice_no': invoice_no,
                })
            else:
                # 保持原样
                all_items.append({
                    'name': name,
                    'spec': spec if spec else '',
                    'unit': unit,
                    'qty': float(qty) if qty is not None else 0,
                    'price': price,
                    'amount': amount_float,
                    'invoice_no': invoice_no,
                })

            total_amount += amount_float

    if file_has_data:
        processed_source_files.append(invoice_file)

if not processed_sheet_names:
    raise ValueError('未找到可汇总的对账数据')

# 按商品名称+规格汇总相同的项
summary_items = defaultdict(lambda: {
    'name': None,
    'spec': None,
    'unit': None,
    'qty': 0,
    'price': None,
    'amount': 0
})

for item in all_items:
    key = (item['name'], item['spec'], item['unit'], item['price'])
    if summary_items[key]['name'] is None:
        summary_items[key]['name'] = item['name']
        summary_items[key]['spec'] = item['spec']
        summary_items[key]['unit'] = item['unit']
        summary_items[key]['price'] = item['price']
    
    summary_items[key]['qty'] += item['qty'] if isinstance(item['qty'], (int, float)) else 0
    summary_items[key]['amount'] += item['amount']

# 转换为列表并按名称排序
summarized = list(summary_items.items())
summarized = sorted(summarized, key=lambda x: (x[0][0], x[0][1], x[0][3]))

# 创建汇总工作簿
summary_wb = openpyxl.Workbook()
summary_ws = summary_wb.active
summary_ws.title = "开票汇总"

# 设置列宽
summary_ws.column_dimensions['A'].width = 18
summary_ws.column_dimensions['B'].width = 20
summary_ws.column_dimensions['C'].width = 12
summary_ws.column_dimensions['D'].width = 12
summary_ws.column_dimensions['E'].width = 15

# 定义样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
title_font = Font(bold=True, size=14)
total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
total_font = Font(bold=True, size=12, color="FF0000")

# 标题行
summary_ws['A1'] = "销售发票"
summary_ws['A1'].font = title_font
summary_ws.merge_cells('A1:E1')
summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
summary_ws.row_dimensions[1].height = 25

# 客户信息行
customer_label = '；'.join(customer_infos) if customer_infos else '客户信息'
address_label = '；'.join(address_infos) if address_infos else '送货地址信息'
source_label = '来源文件：' + '、'.join(processed_source_files)

summary_ws['A2'] = customer_label
summary_ws.merge_cells('A2:E2')
summary_ws['A2'].alignment = Alignment(wrap_text=True)

summary_ws['A3'] = address_label
summary_ws.merge_cells('A3:E3')
summary_ws['A3'].alignment = Alignment(wrap_text=True)

summary_ws['A4'] = source_label
summary_ws.merge_cells('A4:E4')
summary_ws['A4'].alignment = Alignment(wrap_text=True)

# 表头行
row = 6
headers = ['商品名称', '规格型号', '数量', '单价(元)', '金额(元)']
for col, header in enumerate(headers, 1):
    cell = summary_ws.cell(row, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
summary_ws.row_dimensions[row].height = 20

# 数据行
row = 7
for key, item_data in summarized:
    summary_ws.cell(row, 1, item_data['name']).border = border
    summary_ws.cell(row, 2, item_data['spec']).border = border
    
    qty_cell = summary_ws.cell(row, 3)
    qty_cell.value = item_data['qty']
    qty_cell.border = border
    qty_cell.alignment = Alignment(horizontal='right')
    qty_cell.number_format = '0.00'
    
    price_cell = summary_ws.cell(row, 4)
    price_cell.value = item_data['price']
    price_cell.border = border
    price_cell.alignment = Alignment(horizontal='right')
    
    amount_cell = summary_ws.cell(row, 5)
    amount_cell.value = item_data['amount']
    amount_cell.border = border
    amount_cell.alignment = Alignment(horizontal='right')
    amount_cell.number_format = '0.00'
    
    row += 1

# 合计行
summary_ws.cell(row, 1, '合计').font = total_font
summary_ws.cell(row, 1).fill = total_fill
summary_ws.cell(row, 1).border = border

summary_ws.cell(row, 2).fill = total_fill
summary_ws.cell(row, 2).border = border

summary_ws.cell(row, 3).fill = total_fill
summary_ws.cell(row, 3).border = border

summary_ws.cell(row, 4).fill = total_fill
summary_ws.cell(row, 4).border = border

total_cell = summary_ws.cell(row, 5)
total_cell.value = total_amount
total_cell.font = total_font
total_cell.fill = total_fill
total_cell.border = border
total_cell.number_format = '0.00'
total_cell.alignment = Alignment(horizontal='right')

# 保存文件
output_file = f"开票汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
summary_wb.save(output_file)

print(f"✓ 开票汇总已生成: {output_file}")
print(f"\n统计信息:")
print(f"  - 来源文件: {', '.join(processed_source_files)}")
print(f"  - 来源工作表: {', '.join(processed_sheet_names)}")
print(f"  - 商品行数: {len(summarized)}")
print(f"  - 总金额: {total_amount:.2f}元")
print(f"\n商品清单:")
for key, item_data in summarized:
    print(f"  {item_data['name']} ({item_data['spec']}) {item_data['qty']:.2f}{item_data['unit']} @ {item_data['price']}元 = {item_data['amount']:.2f}元")
