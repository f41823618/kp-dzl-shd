"""
根据对账单自动生成送货单工作表
逻辑：
1. 有销售单据号 → 按单据号分组
2. 无单据号但有车号 → 匹配同日期+同车号的已存在单据
3. 无单据号也无车号 → 归入当前分组（通常是橡胶圈等配件）
"""

import openpyxl
from openpyxl import load_workbook
import datetime
import copy
from openpyxl.utils import get_column_letter


def format_date(d):
    """将日期对象格式化为 YYYY年M月D日"""
    if isinstance(d, (datetime.datetime, datetime.date)):
        return f"{d.year}年{d.month}月{d.day}日"
    elif isinstance(d, str):
        try:
            dt = datetime.datetime.strptime(d.strip(), '%Y-%m-%d')
            return f"{dt.year}年{dt.month}月{dt.day}日"
        except Exception:
            return d
    return str(d)


def normalize_car(car):
    """去除车号前后空格及特殊字符"""
    if car is None:
        return None
    return str(car).strip()


# ── 读取对账单 ──────────────────────────────────────────────
dz_wb = load_workbook('对账单.xlsx', data_only=True)
dz_ws = dz_wb.active

# 读取客户/地址
customer_row4 = dz_ws.cell(4, 1).value or ''
address_row5 = dz_ws.cell(5, 1).value or ''
# 从"客户：XXX"中提取名称
customer_name = customer_row4.replace('客户：', '').strip()
address = address_row5.replace('送货地址：', '').strip()

print(f"客户：{customer_name}")
print(f"地址：{address}")

# 解析数据行 (行8~74)
data_rows = []
for i in range(8, 75):
    date = dz_ws.cell(i, 1).value
    name = dz_ws.cell(i, 2).value
    spec = dz_ws.cell(i, 3).value
    unit = dz_ws.cell(i, 4).value
    qty  = dz_ws.cell(i, 5).value
    price= dz_ws.cell(i, 6).value
    car  = normalize_car(dz_ws.cell(i, 8).value)
    order= dz_ws.cell(i, 9).value

    if date is None and name is None:
        continue

    if order:
        order = str(order).strip()

    data_rows.append({
        'row_idx': i,
        'date': date,
        'name': name,
        'spec': spec if spec is not None else '',
        'unit': unit,
        'qty':  qty,
        'price': price,
        'car':  car,
        'order': order,
    })

print(f"\n共读取 {len(data_rows)} 条数据行")

# ── 按送货单分组 ────────────────────────────────────────────
groups = {}          # key → dict{order, date, car, items}
group_order_list = []   # 保持插入顺序
current_group_key = None

for row in data_rows:
    order = row['order']
    car   = row['car']
    date  = row['date']

    if order:
        # 有单据号 → 固定分组
        key = order
        if key not in groups:
            groups[key] = {'order': key, 'date': date, 'car': None, 'items': []}
            group_order_list.append(key)
        current_group_key = key
        # 补充车号
        if car and not groups[key]['car']:
            groups[key]['car'] = car

    elif car:
        # 无单据号，有车号 → 找同日期+同车号
        found_key = None
        for k in group_order_list:
            g = groups[k]
            if g['date'] == date and g['car'] and normalize_car(g['car']) == normalize_car(car):
                found_key = k
                break
        if found_key:
            current_group_key = found_key
        else:
            # 创建新分组（无单号，有车号）
            if isinstance(date, (datetime.datetime, datetime.date)):
                date_str = date.strftime('%Y%m%d')
            else:
                date_str = str(date).replace('-', '')[:8]
            key = f"无单号_{date_str}_{car}"
            if key not in groups:
                groups[key] = {'order': None, 'date': date, 'car': car, 'items': []}
                group_order_list.append(key)
            current_group_key = key
    else:
        # 无单据号，无车号 → 归入当前分组
        pass

    if current_group_key:
        groups[current_group_key]['items'].append(row)

print(f"\n共分为 {len(group_order_list)} 张送货单：")
for k in group_order_list:
    g = groups[k]
    print(f"  {k} | {format_date(g['date'])} | 车号:{g['car']} | {len(g['items'])}条商品")


# ── 构建输出工作簿 ──────────────────────────────────────────
# 以 送货单.xlsx 的 Sheet1 为模板
out_wb = load_workbook('送货单.xlsx')
template_ws = out_wb['Sheet1']

# 删除除 Sheet1 以外的已有工作表
for shname in list(out_wb.sheetnames):
    if shname != 'Sheet1':
        del out_wb[shname]

# 为每个分组创建工作表
for idx, key in enumerate(group_order_list):
    group = groups[key]
    items = group['items']
    date  = group['date']
    car   = group['car'] or ''
    order = group['order'] or ''

    # 复制模板工作表
    new_ws = out_wb.copy_worksheet(template_ws)

    # 工作表命名（最多31字符，去除非法字符）
    sheet_name = order if order else key
    # Excel 工作表名不允许 : / \ ? * [ ]
    for ch in [':', '/', '\\', '?', '*', '[', ']']:
        sheet_name = sheet_name.replace(ch, '_')
    sheet_name = sheet_name[:31]
    # 保证唯一
    base = sheet_name
    suffix = 1
    while sheet_name in out_wb.sheetnames:
        sheet_name = f"{base[:28]}_{suffix}"
        suffix += 1
    new_ws.title = sheet_name

    # ── 填写表头 ──
    new_ws['A3'] = f'送货日期：{format_date(date)}'
    new_ws['E3'] = '单号：'
    new_ws['F3'] = order

    # B4、B5 是合并区域，直接写左上角单元格
    new_ws['A4'] = '客户名称：'
    new_ws['B4'] = customer_name
    new_ws['A5'] = '送货地址：'
    new_ws['B5'] = address

    # ── 清空模板数据行 (7~12) ──
    for r in range(7, 13):
        for c in range(1, 7):
            new_ws.cell(r, c).value = None

    # ── 填写商品明细 ──
    num_items = len(items)

    if num_items <= 6:
        # 正好能放入模板行 7~12
        for i, item in enumerate(items):
            r = 7 + i
            new_ws.cell(r, 1).value = item['name']
            new_ws.cell(r, 2).value = item['spec']
            new_ws.cell(r, 3).value = item['unit']
            new_ws.cell(r, 4).value = item['qty']
            new_ws.cell(r, 5).value = item['price']
            new_ws.cell(r, 6).value = f'=D{r}*E{r}'
        # 合计公式固定覆盖到行12
        total_row = 13
        new_ws[f'F{total_row}'] = '=SUM(F7:F12)'
        car_row = 14
    else:
        # 超过6条，在合计行前插入额外行并调整合计行/车号行
        extra = num_items - 6
        # 在行13之前插入 extra 行
        new_ws.insert_rows(13, extra)
        for i, item in enumerate(items):
            r = 7 + i
            new_ws.cell(r, 1).value = item['name']
            new_ws.cell(r, 2).value = item['spec']
            new_ws.cell(r, 3).value = item['unit']
            new_ws.cell(r, 4).value = item['qty']
            new_ws.cell(r, 5).value = item['price']
            new_ws.cell(r, 6).value = f'=D{r}*E{r}'
        last_data_row = 6 + num_items
        total_row = last_data_row + 1
        new_ws.cell(total_row, 1).value = '合计：'
        new_ws.cell(total_row, 6).value = f'=SUM(F7:F{last_data_row})'
        car_row = total_row + 1
        new_ws.cell(car_row, 1).value = '车号：'
        new_ws.cell(car_row, 4).value = '客户(代表)签收:'

    # ── 填写车号 ──
    new_ws.cell(car_row, 2).value = car

    print(f"  已创建工作表：{sheet_name}")

# ── 将模板工作表（Sheet1）移到最后或删除 ──
# 保留一个空白模板供以后使用，或直接删除原Sheet1
del out_wb['Sheet1']

# ── 保存输出文件 ──
output_path = '送货单_生成.xlsx'
out_wb.save(output_path)
print(f"\n✅ 已保存到：{output_path}")
